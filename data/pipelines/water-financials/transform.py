"""
Water Financials pipeline — transform
Parses Ofwat dividends and long-term cost data into JSON for the site.

Outputs:
  - water-financials.json: industry totals (capex, opex, totex) + dividends over time
  - water-financials-by-company.json: per-company dividends over time
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent.parent.parent
RAW = ROOT / "data" / "raw" / "water-financials"
OUTPUT = ROOT / "data" / "output" / "water-financials"
OUTPUT.mkdir(parents=True, exist_ok=True)

DIVIDENDS_FILE = RAW / "Historic-dividends-since-privatisation.xlsx"
COSTS_FILE = RAW / "Long-term-data-series-v5-Nov-2025-Published.xlsx"

# Company code -> full name mapping
COMPANY_NAMES = {
    "AFW": "Affinity Water",
    "ANH": "Anglian Water",
    "WSH": "Welsh Water (Dwr Cymru)",
    "HDD": "Hafren Dyfrdwy",
    "NES": "Northumbrian Water",
    "PRT": "Portsmouth Water",
    "SVT": "Severn Trent",
    "SEW": "South East Water",
    "SSC": "South Staffs / Cambridge",
    "SBB": "South West Water (Pennon)",
    "SRN": "Southern Water",
    "SES": "SES Water",
    "TMS": "Thames Water",
    "UUW": "United Utilities",
    "WSX": "Wessex Water",
    "YKY": "Yorkshire Water",
}


def parse_dividends():
    """Parse the Appointed Dividend sheet: per-company totals per year."""
    print("  Parsing dividends...")
    wb = openpyxl.load_workbook(str(DIVIDENDS_FILE), data_only=True)
    ws = wb["Appointed Dividend"]

    # Row 4 has year headers starting from col C (index 2)
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    years = []
    for i, v in enumerate(header_row):
        if v and isinstance(v, str) and "-" in v:
            years.append((i, v))  # (column_index, year_label)

    # Find the "Total" rows for each company group
    # Structure: company acronym in col A marks start of group, "Total" in col B marks end
    company_dividends = {}
    current_acronym = None

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]
        col_a = vals[0]
        col_b = str(vals[1]).strip() if vals[1] else ""

        # New company group
        if col_a and isinstance(col_a, str) and col_a.strip() not in ("Sector", "END OF SHEET"):
            current_acronym = col_a.strip()

        # Total row for current company
        if col_b.startswith("Total") and current_acronym:
            year_values = {}
            for col_idx, year_label in years:
                val = vals[col_idx] if col_idx < len(vals) else None
                if val is not None and isinstance(val, (int, float)):
                    year_values[year_label] = round(float(val), 2)
                else:
                    year_values[year_label] = 0.0
            company_dividends[current_acronym] = year_values
            current_acronym = None  # Reset after capturing total

        # Sector total row
        if col_a and str(col_a).strip() == "Sector":
            sector_total = {}
            for col_idx, year_label in years:
                val = vals[col_idx] if col_idx < len(vals) else None
                if val is not None and isinstance(val, (int, float)):
                    sector_total[year_label] = round(float(val), 2)
                else:
                    sector_total[year_label] = 0.0

    wb.close()

    # Build time series
    year_labels = [y[1] for y in years]
    industry_dividends_ts = []
    for year in year_labels:
        industry_dividends_ts.append({
            "year": year,
            "dividends_m": sector_total.get(year, 0.0),
        })

    company_dividends_ts = {}
    for code, year_vals in company_dividends.items():
        name = COMPANY_NAMES.get(code, code)
        series = []
        for year in year_labels:
            series.append({
                "year": year,
                "dividends_m": year_vals.get(year, 0.0),
            })
        company_dividends_ts[code] = {
            "name": name,
            "timeSeries": series,
        }

    print("    Found {} companies, {} years".format(len(company_dividends), len(year_labels)))
    return industry_dividends_ts, company_dividends_ts, year_labels


def parse_costs():
    """Parse the Industry total - real sheet: capex, opex, totex over time."""
    print("  Parsing costs...")
    wb = openpyxl.load_workbook(str(COSTS_FILE), data_only=True)
    ws = wb["Industry total - real"]

    # Row 2 has year headers starting from col K (index 10)
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    years = []
    for i, v in enumerate(header_row):
        if v and isinstance(v, str) and "-" in v:
            years.append((i, v))

    # Row 13: Total capex (water + wastewater)
    # Row 18: Total opex (water + wastewater)
    # Row 23: Total totex (water + wastewater)
    # Also get water/wastewater breakdowns
    rows_data = {}
    target_rows = {
        11: "capex_water",
        12: "capex_wastewater",
        13: "capex_total",
        16: "opex_water",
        17: "opex_wastewater",
        18: "opex_total",
        21: "totex_water",
        22: "totex_wastewater",
        23: "totex_total",
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=False), 1):
        if row_idx in target_rows:
            vals = [c.value for c in row]
            label = target_rows[row_idx]
            row_data = {}
            for col_idx, year_label in years:
                val = vals[col_idx] if col_idx < len(vals) else None
                if val is not None and isinstance(val, (int, float)):
                    row_data[year_label] = round(float(val), 1)
                else:
                    row_data[year_label] = None
            rows_data[label] = row_data

    wb.close()

    # Build time series
    year_labels = [y[1] for y in years]
    costs_ts = []
    for year in year_labels:
        entry = {"year": year}
        for label, year_vals in rows_data.items():
            entry[label + "_m"] = year_vals.get(year)
        costs_ts.append(entry)

    print("    Found {} years of cost data".format(len(year_labels)))
    return costs_ts, year_labels


def main():
    print("=== Water Financials: transform ===")

    if not DIVIDENDS_FILE.exists():
        print("ERROR: Dividends file not found. Run fetch.py first.", file=sys.stderr)
        sys.exit(1)
    if not COSTS_FILE.exists():
        print("ERROR: Costs file not found. Run fetch.py first.", file=sys.stderr)
        sys.exit(1)

    # Parse both datasets
    div_ts, company_div_ts, div_years = parse_dividends()
    costs_ts, cost_years = parse_costs()

    # Build combined industry-level output
    # Merge dividends into cost time series where years overlap
    div_by_year = {d["year"]: d["dividends_m"] for d in div_ts}

    combined_ts = []
    all_years = sorted(set([c["year"] for c in costs_ts] + [d["year"] for d in div_ts]))
    cost_by_year = {}
    for c in costs_ts:
        cost_by_year[c["year"]] = c

    for year in all_years:
        entry = {"year": year}
        # Add cost data if available
        if year in cost_by_year:
            for k, v in cost_by_year[year].items():
                if k != "year":
                    entry[k] = v
        # Add dividend data if available
        if year in div_by_year:
            entry["dividends_m"] = div_by_year[year]
        combined_ts.append(entry)

    # Calculate some summary stats for MetricCards
    # Total dividends since privatisation
    total_dividends = sum(d["dividends_m"] for d in div_ts if d["dividends_m"])
    # Latest year totex
    latest_cost = costs_ts[-1] if costs_ts else {}
    latest_totex = latest_cost.get("totex_total_m")
    latest_capex = latest_cost.get("capex_total_m")

    # Most recent 5 years of dividends
    recent_div = [d for d in div_ts if d["dividends_m"] and d["dividends_m"] > 0]
    recent_5yr_avg = 0.0
    if len(recent_div) >= 5:
        recent_5yr_avg = round(sum(d["dividends_m"] for d in recent_div[-5:]) / 5, 1)

    output = {
        "topic": "water",
        "subtopic": "financials",
        "lastUpdated": "2026-03-03",
        "summary": {
            "totalDividendsSincePrivatisation_m": round(total_dividends, 1),
            "latestTotex_m": latest_totex,
            "latestCapex_m": latest_capex,
            "latestCostYear": costs_ts[-1]["year"] if costs_ts else None,
            "recent5yrAvgDividends_m": recent_5yr_avg,
            "dividendYearRange": "{} to {}".format(div_years[0], div_years[-1]) if div_years else None,
            "costYearRange": "{} to {}".format(cost_years[0], cost_years[-1]) if cost_years else None,
        },
        "national": {
            "timeSeries": combined_ts,
        },
        "byCompany": {
            "dividends": company_div_ts,
        },
        "metadata": {
            "sources": [
                {
                    "name": "Ofwat",
                    "dataset": "Historic dividends since privatisation",
                    "url": "https://www.ofwat.gov.uk/wp-content/uploads/2024/12/Historic-dividends-since-privatisation.xlsx",
                    "retrieved": "2026-03-03",
                    "frequency": "annual",
                    "notes": "Appointed business dividends in nominal GBP millions. Covers 1990-91 to 2023-24."
                },
                {
                    "name": "Ofwat",
                    "dataset": "Long-term data series — company costs (v5, Nov 2025)",
                    "url": "https://www.ofwat.gov.uk/wp-content/uploads/2022/11/Long-term-data-series-v5-Nov-2025-Published.xlsx",
                    "retrieved": "2026-03-03",
                    "frequency": "annual",
                    "notes": "Industry total capex, opex, totex in real terms (2024-25 prices, CPIH-adjusted). Covers 1985-86 to 2024-25."
                }
            ],
            "methodology": "Dividends are nominal (not inflation-adjusted). Costs are real (inflation-adjusted to 2024-25 prices using RPI to 2019-20, then CPIH). Capex includes renewals expenditure for comparability across the totex regime change in 2015. Thames Tideway Tunnel costs included from 2015-16.",
            "knownIssues": [
                "Dividends are nominal; costs are real (inflation-adjusted). Direct comparison requires caution.",
                "Company structures have changed over time through mergers and acquisitions.",
                "Pre-2015 capex/opex split is not directly comparable to post-2015 due to totex regime change. Ofwat has adjusted for consistency.",
                "Some negative dividend values exist (e.g. Anglian 2002-03) representing capital returns or adjustments."
            ]
        }
    }

    # Write main output
    out_file = OUTPUT / "water-financials.json"
    out_file.write_text(json.dumps(output, indent=2))
    print("  Wrote {}".format(out_file))
    print("    {} years of combined data".format(len(combined_ts)))
    print("    Total dividends since privatisation: GBP {:.1f}bn".format(total_dividends / 1000))
    print("    Latest totex ({}): GBP {:.1f}bn".format(
        costs_ts[-1]["year"] if costs_ts else "?",
        latest_totex / 1000 if latest_totex else 0
    ))

    print("Done.")


if __name__ == "__main__":
    main()
